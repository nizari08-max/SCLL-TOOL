"""
output.py — Enriched Excel writer.

Philosophy: the output is the engineer's ORIGINAL workbook, enriched.
We open the original file, preserve every sheet and every cell, then append
FIVE new columns to the right of the line list sheet and add TWO summary
sheets at the end:

    CRITICALITY LEVEL     I / II / III (Roman numerals, colour-coded)
    CLASSIFICATION REASON full text of which rule fired
    CN NUMBER             CN-001… for Level I lines only
    CN REVIEW FLAG        AUTO-CONFIRMED / REVIEW-LARGE-CN / REVIEW-STANDALONE
    DATA QUALITY FLAG     SCOPE: VENDOR | SCOPE: CLIENT | MISSING: … | AMBIGUOUS | blank

Additional sheets appended:
    SUMMARY               classification + CN counts
    CN PROPOSALS          one row per proposed CN with full grouping detail
"""

from __future__ import annotations

from copy import copy
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# Palette
# ─────────────────────────────────────────────────────────────────────────────

COLORS = {
    "LEVEL_I":   "FFFF9999",   # light red
    "LEVEL_II":  "FFFFC266",   # light orange
    "LEVEL_III": "FF92D050",   # light green
    "EXCLUDED":  "FFD9D9D9",   # grey
    "REVIEW":    "FFFFFF99",   # yellow
    "HEADER":    "FF1F4E79",   # dark navy
    "SUMMARY_HDR": "FF2E75B6", # medium blue
    "FLAG_GOOD": "FF92D050",
    "FLAG_WARN": "FFFFC266",
    "FLAG_BAD":  "FFFF9999",
}

FONT_WHITE = Font(color="FFFFFFFF", bold=True)
FONT_BOLD  = Font(bold=True)
THIN = Side(style="thin", color="FF808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NEW_COLUMN_HEADERS = [
    "CRITICALITY LEVEL",
    "CLASSIFICATION REASON",
    "CN NUMBER",
    "CN REVIEW FLAG",
    "DATA QUALITY FLAG",
]

LEVEL_TO_ROMAN = {
    "Level 1": "I",
    "Level 2": "II",
    "Level 3": "III",
}


# ─────────────────────────────────────────────────────────────────────────────
# Public entry
# ─────────────────────────────────────────────────────────────────────────────

def write_enriched_output(
    input_path: str,
    output_path: str,
    detected_config: dict,
    enriched_df: pd.DataFrame,
    cn_proposals: list[dict],
) -> None:
    """
    Enrich the original workbook and save to output_path.

    enriched_df must be the full row-ordered DataFrame (in-scope + excluded
    combined, indexed 0..N-1 matching the original Excel read order) with
    columns Level, Classification_Reason, Data_Quality_Flag, CN_Number,
    CN_Review_Flag already populated.
    """
    wb = openpyxl.load_workbook(input_path)

    sheet_name = detected_config.get("sheet_name")
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    _append_new_columns(ws, detected_config, enriched_df)
    _write_summary_sheet(wb, enriched_df, cn_proposals)
    _write_cn_proposals_sheet(wb, cn_proposals)

    try:
        wb.save(output_path)
    except PermissionError:
        raise PermissionError(
            f"Cannot write to '{output_path}'. "
            f"If the file is open in Excel, please close it first."
        )


# ─────────────────────────────────────────────────────────────────────────────
# 1. Append 5 new columns to the original sheet
# ─────────────────────────────────────────────────────────────────────────────

def _append_new_columns(
    ws, detected_config: dict, enriched_df: pd.DataFrame
) -> None:
    header_row_1based = int(detected_config.get("header_row", 0)) + 1
    skip_rows_0based  = set(detected_config.get("skip_rows", []) or [])

    data_excel_rows = _compute_data_row_indices(ws, header_row_1based, skip_rows_0based)

    # Detect first empty column to the right of existing content
    existing_cols = ws.max_column
    # Try to find the last column that actually has a header value
    last_with_header = existing_cols
    for c in range(existing_cols, 0, -1):
        if ws.cell(row=header_row_1based, column=c).value not in (None, ""):
            last_with_header = c
            break
    first_new_col = last_with_header + 1

    # Header style — mimic header row if possible, else dark navy
    header_fill = _make_fill(COLORS["HEADER"])
    header_template = ws.cell(row=header_row_1based, column=last_with_header)

    for offset, label in enumerate(NEW_COLUMN_HEADERS):
        cell = ws.cell(row=header_row_1based, column=first_new_col + offset, value=label)
        try:
            cell.font = copy(header_template.font) if header_template.font else FONT_WHITE
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill
            cell.font = FONT_WHITE
            cell.border = BORDER
        except Exception:
            cell.fill = header_fill
            cell.font = FONT_WHITE
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = BORDER

    # Write each data row's 5 new values; match by sequential position
    n_rows = min(len(enriched_df), len(data_excel_rows))
    for df_pos in range(n_rows):
        row = enriched_df.iloc[df_pos]
        excel_row = data_excel_rows[df_pos]

        level_raw = str(row.get("Level", "") or "")
        level_roman = LEVEL_TO_ROMAN.get(level_raw, "")
        classification_reason = str(row.get("Classification_Reason", "") or "")
        cn_number = str(row.get("CN_Number", "") or "")
        cn_flag   = str(row.get("CN_Review_Flag", "") or "")
        data_quality = str(row.get("Data_Quality_Flag", "") or "")

        values = [level_roman, classification_reason, cn_number, cn_flag, data_quality]
        row_fill = _row_fill_for(level_raw, data_quality)

        for offset, val in enumerate(values):
            c = ws.cell(row=excel_row, column=first_new_col + offset, value=val)
            if row_fill is not None:
                c.fill = row_fill
            c.border = BORDER
            if offset in (1, 2, 3, 4):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center", bold=False) \
                    if False else Alignment(horizontal="center", vertical="center")

        # Grey out the whole original row if excluded by scope
        if data_quality.startswith("SCOPE:"):
            grey_fill = _make_fill(COLORS["EXCLUDED"])
            for c_idx in range(1, first_new_col + len(NEW_COLUMN_HEADERS)):
                cell = ws.cell(row=excel_row, column=c_idx)
                cell.fill = grey_fill

    # Set widths on the new columns
    widths = [18, 60, 14, 22, 28]
    for offset, w in enumerate(widths):
        letter = get_column_letter(first_new_col + offset)
        ws.column_dimensions[letter].width = w


def _compute_data_row_indices(
    ws, header_row_1based: int, skip_rows_0based: set[int]
) -> list[int]:
    """
    Return the list of openpyxl row numbers that correspond to the pandas
    data rows, matching exactly what pandas read (applies same skip_rows and
    blank-row handling as parser.read_linelist).

    skip_rows_0based are relative to the pandas read (i.e. after header).
    """
    max_col = ws.max_column
    result: list[int] = []
    data_counter = 0  # counts data rows produced by pandas (after skip_rows, before dropna)
    for excel_row in range(header_row_1based + 1, ws.max_row + 1):
        # Is this the units row or an explicit skip?
        # pandas skiprows here are given relative to the original file *by position*,
        # but we conservatively treat skip_rows_0based as the offsets of data rows
        # to skip after the header. Most detectors emit [header_row+1] which means
        # "skip the units row immediately below the header".
        if excel_row - 1 in skip_rows_0based:
            continue
        # Blank row check — pandas dropna(how="all") drops rows that are entirely empty
        row_values = [ws.cell(row=excel_row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_values):
            continue
        result.append(excel_row)
        data_counter += 1
    return result


def _row_fill_for(level: str, data_quality: str) -> Optional[PatternFill]:
    """Pick the cell fill colour for the 5 new cells on a data row."""
    if data_quality.startswith("SCOPE:"):
        return _make_fill(COLORS["EXCLUDED"])
    if data_quality.startswith("MISSING") or data_quality == "AMBIGUOUS":
        return _make_fill(COLORS["REVIEW"])
    if level == "Level 1":
        return _make_fill(COLORS["LEVEL_I"])
    if level == "Level 2":
        return _make_fill(COLORS["LEVEL_II"])
    if level == "Level 3":
        return _make_fill(COLORS["LEVEL_III"])
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 2. SUMMARY sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_summary_sheet(wb, enriched_df: pd.DataFrame, cn_proposals: list[dict]) -> None:
    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]
    ws = wb.create_sheet(title="SUMMARY")

    total = len(enriched_df)
    scope_excluded = int(enriched_df.get("Data_Quality_Flag", pd.Series([], dtype=str))
                         .astype(str).str.startswith("SCOPE:").sum())
    in_scope = total - scope_excluded

    level_series = enriched_df.get("Level", pd.Series([], dtype=str)).astype(str)
    l1 = int((level_series == "Level 1").sum())
    l2 = int((level_series == "Level 2").sum())
    l3 = int((level_series == "Level 3").sum())

    dq_series = enriched_df.get("Data_Quality_Flag", pd.Series([], dtype=str)).astype(str)
    missing = int(dq_series.str.startswith("MISSING").sum())
    ambiguous = int((dq_series == "AMBIGUOUS").sum())

    total_cns = len(cn_proposals)
    auto_cns = sum(1 for p in cn_proposals if p.get("review_flag") == "AUTO-CONFIRMED")
    standalone_cns = sum(1 for p in cn_proposals if p.get("review_flag") == "REVIEW-STANDALONE")
    large_cns = sum(1 for p in cn_proposals if p.get("review_flag") == "REVIEW-LARGE-CN")

    rows = [
        ("SCLL Tool — Classification Summary", None),
        (None, None),
        ("Metric", "Count"),
        ("Total lines in file", total),
        ("Lines in scope", in_scope),
        ("Lines excluded (vendor/client/licensor)", scope_excluded),
        (None, None),
        ("Criticality I (rigorous analysis)", l1),
        ("Criticality II (normal analysis)", l2),
        ("Criticality III (visual check)", l3),
        ("Missing data (cannot classify)", missing),
        ("Ambiguous data", ambiguous),
        (None, None),
        ("CN Summary", None),
        ("Total CNs proposed", total_cns),
        ("CNs auto-confirmed", auto_cns),
        ("CNs flagged REVIEW-LARGE-CN", large_cns),
        ("CNs flagged REVIEW-STANDALONE", standalone_cns),
        (None, None),
        ("Color legend", None),
        ("Criticality I", "Red"),
        ("Criticality II", "Orange"),
        ("Criticality III", "Green"),
        ("Excluded (out of scope)", "Grey"),
        ("Needs review (missing/ambiguous)", "Yellow"),
    ]

    summary_hdr = _make_fill(COLORS["SUMMARY_HDR"])
    legend_fills = {
        "Criticality I":                     _make_fill(COLORS["LEVEL_I"]),
        "Criticality II":                    _make_fill(COLORS["LEVEL_II"]),
        "Criticality III":                   _make_fill(COLORS["LEVEL_III"]),
        "Excluded (out of scope)":           _make_fill(COLORS["EXCLUDED"]),
        "Needs review (missing/ambiguous)":  _make_fill(COLORS["REVIEW"]),
    }

    for r_idx, (label, value) in enumerate(rows, start=1):
        c1 = ws.cell(row=r_idx, column=1, value=label)
        c2 = ws.cell(row=r_idx, column=2, value=value)

        if r_idx == 1:
            c1.font = Font(bold=True, size=14, color="FF1F4E79")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        elif label == "Metric":
            for cc in (c1, c2):
                cc.fill = summary_hdr; cc.font = FONT_WHITE
                cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
        elif label == "CN Summary" or label == "Color legend":
            c1.font = FONT_BOLD
        elif label in legend_fills:
            c1.fill = legend_fills[label]; c2.fill = legend_fills[label]
            c1.border = BORDER; c2.border = BORDER
        elif label and label not in (None,):
            c1.border = BORDER; c2.border = BORDER

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14


# ─────────────────────────────────────────────────────────────────────────────
# 3. CN PROPOSALS sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_cn_proposals_sheet(wb, cn_proposals: list[dict]) -> None:
    if "CN PROPOSALS" in wb.sheetnames:
        del wb["CN PROPOSALS"]
    ws = wb.create_sheet(title="CN PROPOSALS")

    headers = [
        "CN NUMBER", "REVIEW FLAG", "LINES IN CN", "EQUIPMENT TAGS",
        "MIN TEMP (°C)", "MAX TEMP (°C)", "ΔT (°C)",
        "LINE COUNT", "GROUPING REASON", "ENGINEER NOTES",
    ]

    hdr_fill = _make_fill(COLORS["HEADER"])
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill; cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    flag_colors = {
        "AUTO-CONFIRMED":     COLORS["FLAG_GOOD"],
        "REVIEW-LARGE-CN":    COLORS["FLAG_WARN"],
        "REVIEW-STANDALONE":  COLORS["FLAG_BAD"],
    }

    for r_idx, p in enumerate(cn_proposals, start=2):
        flag = p.get("review_flag", "")
        fill = _make_fill(flag_colors.get(flag, "FFFFFFFF"))
        vals = [
            p.get("cn_number", ""),
            flag,
            ", ".join(p.get("line_numbers", []) or []),
            ", ".join(p.get("equipment_tags", []) or []) or "—",
            _num_or_blank(p.get("min_temperature")),
            _num_or_blank(p.get("max_temperature")),
            _num_or_blank(p.get("delta_t")),
            p.get("line_count", ""),
            p.get("grouping_reason", ""),
            "",
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=v)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    widths = [12, 22, 48, 36, 12, 12, 10, 10, 60, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _num_or_blank(val) -> str:
    if val is None:
        return ""
    try:
        return f"{float(val):.0f}"
    except (ValueError, TypeError):
        return str(val)
