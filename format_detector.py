"""
format_detector.py — Format-agnostic auto-detection for any Excel line list.

Detects:
  A) Data sheet         — picks "Line List" or first non-cover sheet
  B) Header row         — scans first 20 rows, picks the one matching most column patterns
  C) Units-row skip     — row after header that is mostly unit strings
  D) Column mappings    — greedy score-based match of actual headers → internal fields
  E) Size unit          — mm vs NPS inches from first 100 non-null size values
  F) Scope strategy     — include_values / text_keywords / column_exclude_values / assume_all
  G) Equipment mode     — tag_prefix (short codes) vs keyword (full descriptive names)

All patterns, scope vocabularies, and equipment keywords live in mapping.yaml —
no project-specific strings are hardcoded in this module.
"""

from __future__ import annotations

import copy
import os
import re
from typing import Optional

import openpyxl
import yaml


# ── Module-level mapping cache ────────────────────────────────────────────────

_MAPPING_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mapping.yaml")
_mapping_cache: dict | None = None


def load_mapping(path: str = _MAPPING_PATH) -> dict:
    """Load mapping.yaml (cached)."""
    global _mapping_cache
    if _mapping_cache is None or path != _MAPPING_PATH:
        with open(path, "r", encoding="utf-8") as f:
            _mapping_cache = yaml.safe_load(f)
    return _mapping_cache


# ── Regex / constants ─────────────────────────────────────────────────────────

UNIT_CELL_RE = re.compile(
    r"^\s*[\(\[]?\s*(?:°c|°f|mm|bar\b|barg|bara|kg/m[23]|psi|in\b|inch|m3/h|m³/h|"
    r"kn|mpa|kpa|%|hrs|days|h\b|kw|mw|rpm|n/a|-+|tbd)\s*[\)\]]?\s*$",
    re.IGNORECASE,
)

NPS_FRACTIONS = {0.5, 0.75, 1.0, 1.25, 1.5, 2.5}

TAG_PREFIX_RE = re.compile(r"^[A-Z]{1,4}-?\d", re.IGNORECASE)

EXACT_BONUS = 50


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def detect_format(filepath: str, mapping: dict | None = None) -> tuple[dict, str]:
    """
    Auto-detect the format of an Excel line list file.
    Returns (detected_config, human_readable_summary).
    """
    if mapping is None:
        mapping = load_mapping()

    field_patterns: dict[str, list[str]] = mapping.get("columns", {})

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        return {}, f"Could not open file: {exc}"

    sheet_names = list(wb.sheetnames)
    data_sheet = _find_data_sheet(sheet_names)

    ws = wb[data_sheet]
    scan_rows = [tuple(row) for row in ws.iter_rows(min_row=1, max_row=20, values_only=True)]
    wb.close()

    warnings: list[str] = []

    header_row_0idx, headers = _detect_header_row(scan_rows, field_patterns, warnings)
    skip_rows = _detect_units_row(scan_rows, header_row_0idx)
    col_mappings, not_found = _detect_column_mappings(headers, field_patterns)

    # Read data rows for further detection (size, scope, equipment)
    wb2 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws2 = wb2[data_sheet]
    skip_1idx = {r + 1 for r in skip_rows}
    header_excel_row = header_row_0idx + 1

    sample_rows: list[tuple] = []
    row_count = 0
    for excel_row, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        if excel_row <= header_excel_row:
            continue
        if excel_row in skip_1idx:
            continue
        if any(c is not None and str(c).strip() not in ("", "nan") for c in row):
            row_count += 1
            if len(sample_rows) < 100:
                sample_rows.append(tuple(row))
    wb2.close()

    # E) Size unit
    size_col_idx = _find_col_idx(headers, col_mappings.get("size", ""))
    size_unit = "inches"
    if size_col_idx is not None:
        size_vals = [r[size_col_idx] for r in sample_rows if size_col_idx < len(r)]
        size_unit = _detect_size_unit(size_vals, warnings)

    # F) Scope strategy
    scope_cfg = _detect_scope(headers, col_mappings, sample_rows, mapping, warnings)

    # G) Equipment mode
    from_col_idx = _find_col_idx(headers, col_mappings.get("from_equipment", ""))
    equip_mode = "tag_prefix"
    if from_col_idx is not None:
        from_vals = [str(r[from_col_idx]) for r in sample_rows
                     if from_col_idx < len(r) and r[from_col_idx] is not None]
        equip_mode = _detect_equipment_mode(from_vals)

    detected_config = {
        "header_row":         header_row_0idx,
        "skip_rows":          skip_rows,
        "sheet_name":         data_sheet,
        "sheet_names":        sheet_names,
        "column_mappings":    col_mappings,
        "not_found_columns":  not_found,
        "size_unit":          size_unit,
        "equipment_mode":     equip_mode,
        "row_count":          row_count,
        "detection_warnings": warnings,
        **scope_cfg,
    }
    return detected_config, _build_summary(detected_config)


# ─────────────────────────────────────────────────────────────────────────────
# A) Data sheet
# ─────────────────────────────────────────────────────────────────────────────

def _find_data_sheet(sheet_names: list[str]) -> str:
    lower = [s.lower() for s in sheet_names]
    for candidate in ("line list", "linelist", "line_list", "lines", "data", "piping"):
        for i, s in enumerate(lower):
            if s == candidate:
                return sheet_names[i]
    non_cover = [s for s in sheet_names if "cover" not in s.lower()]
    return non_cover[0] if non_cover else sheet_names[0]


# ─────────────────────────────────────────────────────────────────────────────
# B) Header row
# ─────────────────────────────────────────────────────────────────────────────

def _detect_header_row(
    scan_rows: list[tuple], field_patterns: dict[str, list[str]], warnings: list
) -> tuple[int, list[str]]:
    best_idx, best_score, best_headers = 0, -1, []
    for idx, row in enumerate(scan_rows[:16]):
        headers = [str(c).strip() if c is not None else "" for c in row]
        score = _score_header_row(headers, field_patterns)
        if score > best_score:
            best_score, best_idx, best_headers = score, idx, headers
    if best_score == 0:
        warnings.append("Header row detection uncertain — using row 1")
    return best_idx, best_headers


def _score_header_row(headers: list[str], field_patterns: dict[str, list[str]]) -> int:
    matched: set[str] = set()
    score = 0
    for cell in headers:
        if not cell:
            continue
        cell_lower = cell.lower()
        for field, patterns in field_patterns.items():
            if field in matched:
                continue
            for pat in patterns:
                if pat in cell_lower:
                    matched.add(field)
                    score += 1
                    break
    non_empty = sum(1 for h in headers if h)
    return score * 10 + non_empty


# ─────────────────────────────────────────────────────────────────────────────
# C) Units row
# ─────────────────────────────────────────────────────────────────────────────

def _detect_units_row(scan_rows: list[tuple], header_row_0idx: int) -> list[int]:
    units_idx = header_row_0idx + 1
    if units_idx >= len(scan_rows):
        return []
    row = scan_rows[units_idx]
    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
    if not cells:
        return []
    unit_hits = sum(1 for c in cells if UNIT_CELL_RE.match(c))
    if unit_hits / len(cells) >= 0.5:
        return [units_idx]
    return []


# ─────────────────────────────────────────────────────────────────────────────
# D) Column mappings (greedy, highest-score-first)
# ─────────────────────────────────────────────────────────────────────────────

def _detect_column_mappings(
    headers: list[str], field_patterns: dict[str, list[str]]
) -> tuple[dict[str, str], list[str]]:
    """
    For each (field, column) pair, compute a score then assign greedily.
    Returns (field → excel_header) and list of fields that could not be mapped.
    """
    candidates: list[tuple[int, str, int]] = []  # (score, field, col_idx)
    for col_idx, header in enumerate(headers):
        if not header:
            continue
        h = header.lower()
        for field, patterns in field_patterns.items():
            best = 0
            for pat in patterns:
                if h == pat:
                    best = max(best, EXACT_BONUS + 100)
                elif h.startswith(pat) and len(pat) >= 3:
                    best = max(best, 80 + EXACT_BONUS - len(h))
                elif len(pat) >= 4 and pat in h:
                    best = max(best, 40 + len(pat) - len(h) // 2)
            if best > 0:
                candidates.append((best, field, col_idx))
    candidates.sort(reverse=True)

    field_to_col: dict[str, int] = {}
    used_cols: set[int] = set()
    for _, field, col_idx in candidates:
        if field in field_to_col or col_idx in used_cols:
            continue
        field_to_col[field] = col_idx
        used_cols.add(col_idx)

    col_mappings = {f: headers[c] for f, c in field_to_col.items()}

    # Core fields the classifier benefits from (used for "not found" reporting)
    core = [
        "line_number", "size", "design_temperature", "material",
        "from_equipment", "to_equipment", "inlet_pressure", "fluid_service",
        "relief_line", "vibration", "expansion_joint", "vacuum", "settlement",
        "jacketed", "cement_lined", "vertical_tower", "ped_category_3",
        "nozzle_load_limit", "heavy_wall", "cyclic_service", "category_m",
        "client_request", "underground", "schedule_160",
    ]
    not_found = [f for f in core if f not in col_mappings]
    return col_mappings, not_found


# ─────────────────────────────────────────────────────────────────────────────
# E) Size unit
# ─────────────────────────────────────────────────────────────────────────────

def _detect_size_unit(size_vals: list, warnings: list) -> str:
    numeric = []
    for v in size_vals[:30]:
        if v is None:
            continue
        try:
            numeric.append(float(str(v).strip()))
        except (ValueError, TypeError):
            continue
    if not numeric:
        return "inches"
    has_large = any(v >= 15 for v in numeric)
    has_nps_fractions = any(v in NPS_FRACTIONS for v in numeric)
    if has_large and not has_nps_fractions:
        return "mm"
    if has_nps_fractions:
        return "inches"
    if has_large:
        warnings.append("Size unit ambiguous (no NPS fractions, large values) — assuming mm")
        return "mm"
    return "inches"


# ─────────────────────────────────────────────────────────────────────────────
# F) Scope strategy
# ─────────────────────────────────────────────────────────────────────────────

def _detect_scope(
    headers: list[str],
    col_mappings: dict[str, str],
    sample_rows: list[tuple],
    mapping: dict,
    warnings: list,
) -> dict:
    scope_vocab = mapping.get("scope_values", {})
    include_vocab = {v.strip().lower() for v in scope_vocab.get("include", [])}
    exclude_vocab = {v.strip().lower() for v in scope_vocab.get("exclude", [])}
    text_keywords = [k.strip().lower() for k in mapping.get("scope_text_keywords", [])]

    # 1) Dedicated scope column
    scope_idx = _find_col_idx(headers, col_mappings.get("scope", ""))
    if scope_idx is not None:
        vals = [str(r[scope_idx]).strip().lower() for r in sample_rows
                if scope_idx < len(r) and r[scope_idx] is not None
                and str(r[scope_idx]).strip() not in ("", "nan")]
        if vals and (any(v in include_vocab for v in vals) or any(v in exclude_vocab for v in vals)):
            return {
                "scope_mode":         "include_values",
                "scope_column":       col_mappings["scope"],
                "scope_include_values": sorted(include_vocab),
                "scope_exclude_values": [],
                "scope_exclude_keywords": [],
            }

    # 2) Notes / remarks column with vendor/client text
    notes_col = col_mappings.get("notes_remarks", "")
    notes_idx = _find_col_idx(headers, notes_col)
    if notes_idx is not None and text_keywords:
        vals = [str(r[notes_idx]).strip().lower() for r in sample_rows
                if notes_idx < len(r) and r[notes_idx] is not None]
        if any(kw in v for v in vals for kw in text_keywords):
            return {
                "scope_mode":           "text_keywords",
                "scope_column":         notes_col,
                "scope_include_values": [],
                "scope_exclude_values": [],
                "scope_exclude_keywords": text_keywords,
            }

    # 3) Material column contains "not in scope" values
    mat_col = col_mappings.get("material", "")
    mat_idx = _find_col_idx(headers, mat_col)
    if mat_idx is not None:
        exclude_hits = [str(r[mat_idx]).strip() for r in sample_rows
                        if mat_idx < len(r) and r[mat_idx] is not None
                        and str(r[mat_idx]).strip().lower() in exclude_vocab]
        if exclude_hits:
            return {
                "scope_mode":           "column_exclude_values",
                "scope_column":         mat_col,
                "scope_include_values": [],
                "scope_exclude_values": sorted(set(exclude_hits)),
                "scope_exclude_keywords": [],
            }

    warnings.append("No scope column detected — all lines treated as in-scope")
    return {
        "scope_mode":           "assume_all_in_scope",
        "scope_column":         "",
        "scope_include_values": [],
        "scope_exclude_values": [],
        "scope_exclude_keywords": [],
    }


# ─────────────────────────────────────────────────────────────────────────────
# G) Equipment mode
# ─────────────────────────────────────────────────────────────────────────────

def _detect_equipment_mode(from_vals: list[str]) -> str:
    tag_count, keyword_count = 0, 0
    for v in from_vals[:30]:
        if not v or v.strip().lower() in ("", "nan", "none"):
            continue
        for seg in (s.strip() for s in re.split(r"[/&;,]", v) if s.strip()):
            if " " in seg:
                keyword_count += 1
                break
            if TAG_PREFIX_RE.match(seg) and len(seg) <= 12:
                tag_count += 1
    return "keyword" if keyword_count >= tag_count else "tag_prefix"


# ─────────────────────────────────────────────────────────────────────────────
# Merge detection into rules
# ─────────────────────────────────────────────────────────────────────────────

def apply_detection_to_rules(
    rules: dict, detected_config: dict, mapping: dict | None = None
) -> dict:
    """
    Return a deep copy of rules with detection-driven fields injected:
      - strain_sensitive_equipment.detection_mode (from detected_config)
      - strain_sensitive_equipment.keyword_patterns (from mapping.yaml)
      - strain_sensitive_equipment.non_strain_sensitive_keywords (from mapping.yaml)
      - size_config.unit (from detected_config)

    Parser and classifier read ONLY from this merged rules dict.
    """
    if mapping is None:
        mapping = load_mapping()
    merged = copy.deepcopy(rules)

    sse = merged.setdefault("strain_sensitive_equipment", {})
    sse["detection_mode"] = detected_config.get("equipment_mode", "tag_prefix")

    kw_cfg = mapping.get("equipment_keywords", {})
    if "strain_sensitive" in kw_cfg:
        sse["keyword_patterns"] = [
            {"type": e["type"], "keywords": list(e["keywords"])}
            for e in kw_cfg["strain_sensitive"]
        ]
    if "non_strain_sensitive" in kw_cfg:
        sse["non_strain_sensitive_keywords"] = list(kw_cfg["non_strain_sensitive"])

    size_cfg = merged.setdefault("size_config", {})
    size_cfg["unit"] = detected_config.get("size_unit", "inches")

    return merged


# ─────────────────────────────────────────────────────────────────────────────
# Summary builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary(cfg: dict) -> str:
    out = []
    out.append(f"Header row: Excel row {cfg['header_row'] + 1}")
    if cfg.get("skip_rows"):
        out.append(f"Skipped rows: {[r + 1 for r in cfg['skip_rows']]} (units row)")
    out.append(f"Size unit: {cfg['size_unit']}"
               + (" → converted to NPS inches" if cfg['size_unit'] == "mm" else ""))

    scope_mode = cfg.get("scope_mode", "assume_all_in_scope")
    scope_col = cfg.get("scope_column", "")
    scope_labels = {
        "include_values":         f"dedicated scope column '{scope_col}'",
        "text_keywords":          f"keyword detection in '{scope_col}'",
        "column_exclude_values":  f"exclude values in '{scope_col}'",
        "assume_all_in_scope":    "no scope column — all lines in scope",
    }
    out.append(f"Scope: {scope_labels.get(scope_mode, scope_mode)}")
    out.append(f"Equipment: {'keyword mode' if cfg['equipment_mode'] == 'keyword' else 'tag-prefix mode'}")

    mapped = len(cfg.get("column_mappings", {}))
    not_found = cfg.get("not_found_columns", [])
    out.append(f"Columns mapped: {mapped} (not found: {len(not_found)})")
    if not_found:
        out.append("  Missing: " + ", ".join(not_found))

    out.append(f"Data rows: {cfg.get('row_count', '?')}")
    for w in cfg.get("detection_warnings", []):
        out.append(f"! {w}")
    return "\n".join(out)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _find_col_idx(headers: list[str], target: str) -> Optional[int]:
    if not target:
        return None
    t = target.strip().lower()
    for i, h in enumerate(headers):
        if str(h).strip().lower() == t:
            return i
    return None
