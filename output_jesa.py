"""
output_jesa.py — JESA SCLL format Excel writer.

Produces an output workbook that preserves the original JESA line list column
structure and enriches it with four new tool-generated columns:

  CLASSIFICATION REASON  — inserted after STRESS CRITICALITY column
  SCOPE STATUS           — inserted after CLASSIFICATION REASON
  REVIEW FLAG            — inserted after SCOPE STATUS
  CN REVIEW FLAG         — inserted after CALCULATION NUMBER column

Also writes a SUMMARY sheet with project info, level counts, and CN table.

No classification logic here. Receives already-classified DataFrames.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import collections


# ─────────────────────────────────────────────────────────────────────────────
# Color palette (ARGB format for openpyxl) — matches JESA SCLL convention
# ─────────────────────────────────────────────────────────────────────────────
COLORS = {
    "I":            "FFFF0000",   # Level I   — red
    "II":           "FFFFA500",   # Level II  — orange
    "III":          "FF00B050",   # Level III — green
    "EXCLUDED":     "FFC0C0C0",   # excluded  — grey
    "NEEDS REVIEW": "FFFFFF00",   # review    — yellow
    "DASH":         "FFC0C0C0",   # dash / not assessed — grey
    "HEADER":       "FF1F4E79",   # dark blue header
    "SUMMARY_HDR":  "FF2E75B6",   # medium blue
    "NEW_COL":      "FFE2EFDA",   # light green tint for new tool columns header
}

LEVEL_TO_ROMAN = {
    "Level 1": "I",
    "Level 2": "II",
    "Level 3": "III",
    "":        "-",
}

FONT_WHITE = Font(color="FFFFFFFF", bold=True)
FONT_BOLD  = Font(bold=True)
FONT_DARK  = Font(color="FF000000")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _row_fill_key(level_roman: str, review_flag: str, is_excluded: bool) -> str:
    if is_excluded:
        return "EXCLUDED"
    if review_flag == "NEEDS REVIEW":
        return "NEEDS REVIEW"
    if level_roman in ("I", "II", "III"):
        return level_roman
    return "DASH"


def _auto_size_columns(ws, max_width: int = 45) -> None:
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ─────────────────────────────────────────────────────────────────────────────
# Build the merged column list for output
# ─────────────────────────────────────────────────────────────────────────────

def _build_output_columns(classified_df: pd.DataFrame, rules: dict) -> list[tuple[str, str]]:
    """
    Build the ordered list of (internal_name, display_header) pairs for output.

    Original DataFrame columns are emitted in their natural order (as read from
    the input Excel).  After the stress_criticality column the three new analysis
    columns are injected; after the calculation_number column the CN review flag
    is injected.  The internal Level/Review_Flag/Classification_Reason/CN columns
    are NOT emitted as separate columns — their values are used to populate the
    new injected columns and to overwrite stress_criticality.
    """
    col_map = rules.get("column_mappings", {})
    # Build reverse: internal_name → display_header
    reverse: dict[str, str] = {v: k for k, v in col_map.items() if v}
    # Also build: internal_name → original Excel header (for display)
    internal_to_display: dict[str, str] = {}
    for internal, excel_hdr in col_map.items():
        if excel_hdr:
            internal_to_display[internal] = excel_hdr

    # Columns added by classifier/cn_assigner — kept separate, not emitted
    tool_cols = {"Level", "Classification_Reason", "Review_Flag",
                 "Proposed_CN", "CN_Reason", "CN_Review_Flag", "CN_Status",
                 "Scope_Filter_Result"}

    stress_col = "stress_criticality"
    calc_col   = "calculation_number"

    new_analysis_cols = [
        ("_CLASSIFICATION_REASON", "CLASSIFICATION REASON"),
        ("_SCOPE_STATUS",          "SCOPE STATUS"),
        ("_REVIEW_FLAG",           "REVIEW FLAG"),
    ]
    new_cn_flag_col = ("_CN_REVIEW_FLAG", "CN REVIEW FLAG")

    result: list[tuple[str, str]] = []
    for col in classified_df.columns:
        if col in tool_cols:
            continue
        display = internal_to_display.get(col, col)
        result.append((col, display))
        if col == stress_col:
            result.extend(new_analysis_cols)
        if col == calc_col:
            result.append(new_cn_flag_col)

    # If stress_criticality or calculation_number not found in df, append at end
    injected_keys = {c[0] for c in new_analysis_cols + [new_cn_flag_col]}
    already = {c[0] for c in result}
    for item in new_analysis_cols + [new_cn_flag_col]:
        if item[0] not in already:
            result.append(item)

    return result


def _get_cell_value(row: pd.Series, col_internal: str, is_excluded: bool,
                    cn_proposals_by_line: dict) -> object:
    """Return the value to write for a given internal column name."""
    # Tool-injected columns (prefixed with _)
    if col_internal == "_CLASSIFICATION_REASON":
        return row.get("Classification_Reason", "")

    if col_internal == "_SCOPE_STATUS":
        return "EXCLUDED" if is_excluded else "STRESS_SCOPE"

    if col_internal == "_REVIEW_FLAG":
        rv = str(row.get("Review_Flag", "")).strip()
        if rv in ("", "nan"):
            return "OK"
        return rv

    if col_internal == "_CN_REVIEW_FLAG":
        flag = str(row.get("CN_Review_Flag", "")).strip()
        if flag in ("", "nan"):
            return ""
        return flag

    # stress_criticality: overwrite with Roman numeral from computed Level
    if col_internal == "stress_criticality":
        level = str(row.get("Level", "")).strip()
        return LEVEL_TO_ROMAN.get(level, "-")

    # calculation_number: use Proposed_CN if available, else keep original
    if col_internal == "calculation_number":
        proposed = str(row.get("Proposed_CN", "")).strip()
        original = str(row.get("calculation_number", "")).strip()
        if proposed and proposed.lower() not in ("", "nan"):
            return proposed
        if original and original.lower() not in ("", "nan"):
            return original
        return ""

    val = row.get(col_internal, "")
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if str(val).strip().lower() in ("nan", "none"):
        return ""
    return val


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Line List
# ─────────────────────────────────────────────────────────────────────────────

def write_line_list_sheet(
    wb: openpyxl.Workbook,
    classified_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    rules: dict,
    cn_proposals: list,
) -> None:
    ws = wb.active
    ws.title = "Line List"

    # Tag excluded rows
    classified_with_flag = classified_df.copy()
    classified_with_flag["_is_excluded"] = False

    excluded_with_flag = excluded_df.copy()
    excluded_with_flag["_is_excluded"] = True
    excluded_with_flag["Level"] = ""
    excluded_with_flag["Classification_Reason"] = "Line excluded from stress scope"
    excluded_with_flag["Review_Flag"] = "EXCLUDED"
    for col in ("Proposed_CN", "CN_Reason", "CN_Review_Flag", "CN_Status"):
        if col not in excluded_with_flag.columns:
            excluded_with_flag[col] = ""

    combined = pd.concat([classified_with_flag, excluded_with_flag], ignore_index=True)

    # Build CN lookup by line number (for quick access)
    cn_by_line: dict = {}
    for prop in cn_proposals:
        for ln in prop.get("line_numbers", []):
            cn_by_line[str(ln)] = prop

    output_cols = _build_output_columns(combined, rules)

    # ── Header row ────────────────────────────────────────────────────────────
    new_col_keys = {"_CLASSIFICATION_REASON", "_SCOPE_STATUS",
                    "_REVIEW_FLAG", "_CN_REVIEW_FLAG"}
    header_fill     = _make_fill(COLORS["HEADER"])
    new_col_fill    = _make_fill(COLORS["NEW_COL"])

    for col_idx, (internal, display) in enumerate(output_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display)
        if internal in new_col_keys:
            cell.fill = new_col_fill
            cell.font = Font(bold=True, color="FF1F4E79")
        else:
            cell.fill = header_fill
            cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_pos, (_, data_row) in enumerate(combined.iterrows(), start=2):
        is_excluded  = bool(data_row.get("_is_excluded", False))
        level_val    = str(data_row.get("Level", "")).strip()
        roman_level  = LEVEL_TO_ROMAN.get(level_val, "-")
        review_flag  = str(data_row.get("Review_Flag", "")).strip()

        fill_key = _row_fill_key(roman_level, review_flag, is_excluded)
        fill     = _make_fill(COLORS[fill_key])

        for col_idx, (internal, display) in enumerate(output_cols, start=1):
            cell = ws.cell(row=row_pos, column=col_idx)
            cell.value  = _get_cell_value(data_row, internal, is_excluded, cn_by_line)
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=False, vertical="center")

    ws.freeze_panes = "A2"

    # Fixed widths for key columns
    ws.row_dimensions[1].height = 40

    # Auto-size, then cap long columns
    for col_idx, (internal, display) in enumerate(output_cols, start=1):
        col_letter = get_column_letter(col_idx)
        if internal in new_col_keys:
            ws.column_dimensions[col_letter].width = 35 if internal == "_CLASSIFICATION_REASON" else 18
        else:
            # Simple width based on header length
            ws.column_dimensions[col_letter].width = min(max(len(display) + 2, 8), 30)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def write_summary_sheet(
    wb: openpyxl.Workbook,
    classified_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    rules: dict,
    cn_proposals: list,
) -> None:
    ws = wb.create_sheet(title="SUMMARY")
    out_cfg = rules.get("output_config", {})

    project_name = out_cfg.get("project_name", "")
    project_no   = out_cfg.get("project_number", "")
    customer     = out_cfg.get("customer", "")
    doc_no       = out_cfg.get("document_number", "")

    total_all      = len(classified_df) + len(excluded_df)
    total_excluded = len(excluded_df)
    total_scope    = len(classified_df)

    level_i   = int((classified_df.get("Level", pd.Series()) == "Level 1").sum())
    level_ii  = int((classified_df.get("Level", pd.Series()) == "Level 2").sum())
    level_iii = int((classified_df.get("Level", pd.Series()) == "Level 3").sum())
    review    = int((classified_df.get("Review_Flag", pd.Series()) == "NEEDS REVIEW").sum())
    not_assessed = total_scope - level_i - level_ii - level_iii - review

    total_cns = len(cn_proposals)

    # CN breakdown by material
    moc_col = "material"
    cn_moc_counts: dict = collections.Counter()
    for prop in cn_proposals:
        for li in prop.get("line_df_indices", []):
            try:
                moc = str(classified_df.at[li, moc_col]).strip() if li in classified_df.index else ""
            except Exception:
                moc = ""
            group_moc = _moc_to_group(moc)
            cn_moc_counts[group_moc] += 1
            break  # one MOC label per CN (use first line's MOC)

    hdr_fill    = _make_fill(COLORS["SUMMARY_HDR"])
    l1_fill     = _make_fill(COLORS["I"])
    l2_fill     = _make_fill(COLORS["II"])
    l3_fill     = _make_fill(COLORS["III"])
    excl_fill   = _make_fill(COLORS["EXCLUDED"])
    review_fill = _make_fill(COLORS["NEEDS REVIEW"])

    def _write_row(r, label, value, fill=None, bold=False, big=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        for c in (c1, c2):
            if fill:
                c.fill = fill
            if bold or big:
                c.font = Font(bold=True, size=13 if big else 11)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    row = 1
    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    title_cell = ws.cell(row=row, column=1, value="SCLL TOOL — CLASSIFICATION SUMMARY")
    title_cell.font = Font(bold=True, size=14, color="FF1F4E79")
    row += 1

    # Project info
    ws.cell(row=row, column=1, value="Project:").font = FONT_BOLD
    ws.cell(row=row, column=2, value=project_name)
    row += 1
    ws.cell(row=row, column=1, value="Project No.:").font = FONT_BOLD
    ws.cell(row=row, column=2, value=project_no)
    row += 1
    ws.cell(row=row, column=1, value="Customer:").font = FONT_BOLD
    ws.cell(row=row, column=2, value=customer)
    row += 1
    ws.cell(row=row, column=1, value="Document No.:").font = FONT_BOLD
    ws.cell(row=row, column=2, value=doc_no)
    row += 2

    # Header
    for col_idx, hdr in enumerate(["Metric", "Count"], start=1):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        c.fill = hdr_fill
        c.font = FONT_WHITE
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    row += 1

    _write_row(row, "Total lines in file",                   total_all);    row += 1
    _write_row(row, "Lines in stress scope (classified)",    total_scope);  row += 1
    _write_row(row, "Lines excluded from stress scope",      total_excluded, fill=excl_fill); row += 1
    row += 1

    _write_row(row, "Level I   — Rigorous analysis (Caesar II)", level_i,  fill=l1_fill,   bold=True); row += 1
    _write_row(row, "Level II  — Normal analysis (manual calc)",  level_ii, fill=l2_fill,   bold=True); row += 1
    _write_row(row, "Level III — Visual / approximate check",     level_iii,fill=l3_fill,   bold=True); row += 1
    _write_row(row, "NEEDS REVIEW (missing data — TBD)",          review,   fill=review_fill); row += 1
    _write_row(row, "Not assessed (dash) — engineer decision needed", max(0, not_assessed),
               fill=excl_fill); row += 1
    row += 1

    _write_row(row, "Total proposed CNs (Level I only)", total_cns, bold=True); row += 1

    # CN breakdown by material
    for moc_grp in ("HDPE/FRP", "CS", "SS", "Mixed"):
        count = sum(v for k, v in cn_moc_counts.items() if k == moc_grp)
        _write_row(row, f"  CNs — {moc_grp}", count); row += 1
    row += 1

    # CN breakdown by fluid service (top 10)
    fluid_col = "fluid_service_code"
    cn_fluid_counts: dict = collections.Counter()
    for prop in cn_proposals:
        for li in prop.get("line_df_indices", []):
            try:
                fluid = str(classified_df.at[li, fluid_col]).strip() if li in classified_df.index else ""
            except Exception:
                fluid = ""
            if fluid and fluid.lower() not in ("", "nan"):
                cn_fluid_counts[fluid] += 1
                break

    ws.cell(row=row, column=1, value="CN breakdown by fluid service (top 10)").font = FONT_BOLD
    row += 1
    for fluid, count in cn_fluid_counts.most_common(10):
        _write_row(row, f"  {fluid}", count); row += 1
    row += 1

    # CN detail table header
    cn_table_headers = [
        "CN Number", "Lines", "FROM Equipment", "TO Equipment",
        "Material", "Temp Range (°C)", "Review Flag"
    ]
    for col_idx, hdr in enumerate(cn_table_headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        c.fill = hdr_fill
        c.font = FONT_WHITE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER
    row += 1

    flag_colors = {
        "[AUTO-CONFIRMED]":      COLORS["III"],
        "[REVIEW-TEMPERATURE]":  COLORS["NEEDS REVIEW"],
        "[REVIEW-MODEL-SIZE]":   COLORS["II"],
        "[REVIEW-MISSING-DATA]": COLORS["I"],
        "[REVIEW-AREA-CONFLICT]":COLORS["II"],
        "[REVIEW-MANUAL]":       COLORS["NEEDS REVIEW"],
    }

    for prop in sorted(cn_proposals, key=lambda p: p["cn_number"]):
        flag = prop.get("review_flag", "")
        row_fill = _make_fill(flag_colors.get(flag, "FFFFFFFF"))

        cn_num  = prop.get("cn_number", "")
        n_lines = prop.get("line_count", "")
        equip   = ", ".join(prop.get("equipment_tags", []))[:80]
        min_t   = prop.get("min_temperature", "")
        max_t   = prop.get("max_temperature", "")
        temp_rng = f"{min_t}–{max_t}" if min_t is not None and max_t is not None else "—"

        # Derive material from first line
        first_li = prop.get("line_df_indices", [None])[0]
        try:
            mat = str(classified_df.at[first_li, moc_col]).strip() if first_li in classified_df.index else ""
        except Exception:
            mat = ""

        # Split equip into FROM/TO heuristic
        equip_tokens = equip.split(",")
        frm_eq = equip_tokens[0].strip() if equip_tokens else ""
        to_eq  = equip_tokens[-1].strip() if len(equip_tokens) > 1 else ""

        vals = [cn_num, n_lines, frm_eq, to_eq, mat, temp_rng, flag]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_idx, value=v if v is not None else "")
            c.fill = row_fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=False)
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 25


def _moc_to_group(moc: str) -> str:
    moc_upper = moc.upper()
    if moc_upper in ("HDPE", "FRP"):
        return "HDPE/FRP"
    if moc_upper in ("CS", "CSG", "CS300", "RLCS"):
        return "CS"
    if moc_upper in ("SS", "PTFECS"):
        return "SS"
    if moc_upper:
        return "Mixed"
    return "Mixed"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3 — CN Proposals (reuse output.py style)
# ─────────────────────────────────────────────────────────────────────────────

def write_cn_proposals_sheet(wb: openpyxl.Workbook, cn_proposals: list) -> None:
    ws = wb.create_sheet(title="CN Proposals")

    headers = [
        "CN Number", "Review Flag", "Lines in CN", "Equipment Tags",
        "Area / Unit", "Min Temp (°C)", "Max Temp (°C)", "Delta T (°C)",
        "Line Count", "Grouping Reason",
    ]
    flag_colors = {
        "[AUTO-CONFIRMED]":      "FF92D050",
        "[REVIEW-TEMPERATURE]":  "FFFFFF99",
        "[REVIEW-MODEL-SIZE]":   "FFFFC266",
        "[REVIEW-MISSING-DATA]": "FFFF9999",
        "[REVIEW-AREA-CONFLICT]":"FFFFC266",
        "[REVIEW-MANUAL]":       "FFFFFF99",
    }

    header_fill = _make_fill(COLORS["HEADER"])
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.fill = header_fill
        c.font = FONT_WHITE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER

    for row_idx, prop in enumerate(sorted(cn_proposals, key=lambda p: p["cn_number"]), start=2):
        flag  = prop.get("review_flag", "")
        fill  = _make_fill(flag_colors.get(flag, "FFFFFFFF"))
        vals  = [
            prop.get("cn_number", ""),
            flag,
            ", ".join(prop.get("line_numbers", [])),
            ", ".join(prop.get("equipment_tags", [])),
            ", ".join(prop.get("area_codes", [])) or "—",
            prop.get("min_temperature", ""),
            prop.get("max_temperature", ""),
            prop.get("delta_t", ""),
            prop.get("line_count", ""),
            prop.get("grouping_reason", ""),
        ]
        for col_idx, v in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = "" if v is None else v
            c.fill  = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(wrap_text=(col_idx == 10))

    ws.freeze_panes = "A2"
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["J"].width = 70


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def write_jesa_output(
    classified_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    output_path: str,
    rules: dict,
    cn_proposals: list | None = None,
) -> None:
    """
    Write the JESA SCLL output workbook.

    Sheets:
      1. Line List  — original columns + 4 new tool columns, color-coded
      2. SUMMARY    — project info, level counts, CN table
      3. CN Proposals — one row per proposed CN (if cn_proposals non-empty)
    """
    if cn_proposals is None:
        cn_proposals = []

    wb = openpyxl.Workbook()
    write_line_list_sheet(wb, classified_df, excluded_df, rules, cn_proposals)
    write_summary_sheet(wb, classified_df, excluded_df, rules, cn_proposals)
    if cn_proposals:
        write_cn_proposals_sheet(wb, cn_proposals)

    try:
        wb.save(output_path)
    except PermissionError:
        raise PermissionError(
            f"Cannot write to '{output_path}'. "
            f"If the file is open in Excel, please close it first."
        )
